import json
import openpyxl
import re
from tkinter import filedialog


class Assistant:
    def __init__(self, jsonFileName='Beams_info', xlsxFilePath=None):
        self.fileName = jsonFileName
        if xlsxFilePath is None:
            xlsxFilePath = filedialog.askopenfilename(filetypes=(("Excel files", "*xlsx"), ("Excel files", "*xlsm")))
        self.wb = openpyxl.load_workbook(xlsxFilePath, read_only=True, data_only=True, keep_vba=False)

    def get_variable_value(self, variable):
        with open(self.fileName) as jsonFile:
            dictionary = json.load(jsonFile)
            value = dictionary[variable]
        return value

    def set_variable_value(self, variable, value):
        with open(self.fileName) as jsonFile:
            dictionary = json.load(jsonFile)
        with open(self.fileName, 'w') as jsonFile:
            dictionary[variable] = value
            json.dump(dictionary, jsonFile)

    def set_default_variable(self, key, value):
        with open(self.fileName) as jsonFile:
            dictionary = json.load(jsonFile)
        dictionary.setdefault(str(key), value)
        with open(self.fileName, 'w') as jsonFile:
            json.dump(dictionary, jsonFile)

    def reset_values(self):
        default_values = {
            "Beams_info": {}
        }
        dictionary = default_values.get(self.fileName)
        with open(self.fileName, 'w') as jsonFile:
            json.dump(dictionary, jsonFile)

    def download_excel_beams_info(self, star_index: int = 6, last_index: int = None):
        storey_beam_numbers = {}
        self.reset_values()
        for index in range(star_index, (len(self.wb.sheetnames) if last_index is None else last_index)):
            span_info = self.download_excel_span_info(index)
            beam_storey = re.findall('V(.+?)-', span_info['span_name'])[0]
            try:
                beam_number = re.findall("-(.+?)\(", span_info['span_name'])[0]
            except BaseException:
                beam_number = span_info['span_name'].split('-')[1]
            beam_name = 'V' + beam_storey + '-' + beam_number
            if beam_storey not in storey_beam_numbers:
                storey_beam_numbers.setdefault(beam_storey, [])
            if beam_number not in storey_beam_numbers[beam_storey]:
                storey_beam_numbers[beam_storey].append(beam_number)
            self.set_default_variable(beam_name, {
                "beam_name": beam_name,
                "spans_num": 0,
                "spans_info": []
            })
            beam_dict = self.get_variable_value(beam_name)
            spans_num = beam_dict['spans_num']
            spans_info = beam_dict['spans_info']
            spans_info.append(span_info)
            beam_dict.update({
                "beam_name": beam_name,
                "spans_num": spans_num + 1,
                "spans_info": spans_info
            })
            self.set_variable_value(beam_name, beam_dict)

    def download_excel_span_info(self, index):
        ws = self.wb[self.wb.sheetnames[index]]
        span_keys = ['span_name', 'left_support_info', 'right_support_info', 'free_length', 'width', 'height',
                     'bars_info', 'stirrups_info']
        span_name = ws.title
        ls_info = [ws['L78'].value, ws['N78'].value]
        rs_info = [ws['L80'].value, ws['N80'].value]
        free_length = ws['L79'].value
        width = ws['Q78'].value / 100
        height = ws['Q79'].value / 100

        bars_info = {'quantity': {
            'top_left': 0,
            'top_right': 0,
            'bottom_left': 0,
            'bottom_center': 0,
            'bottom_right': 0},
            'info': []}
        bars_keys = ['label', 'case', 'side', 'order', 'left_cut', 'right_cut', 'tie_info']
        # Top long bar
        if ws['O90'].value != 0 or ws['O91'].value != 0:
            label = ""
            label_1 = str(ws['O90'].value) + '%%C' + ws['Q90'].value
            label_2 = str(ws['O91'].value) + '%%C' + ws['Q91'].value
            if ws['O90'].value != 0:
                label = label + label_1 + (' + ' if ws['O91'].value != 0 else '')
            if ws['O91'].value != 0:
                label = label + label_2
            case = 0
            side = 1
            order = 0
            left_cut = [ws['F189'].value - ws['F219'].value,
                        ws['F190'].value - ws['F219'].value]
            right_cut = [ws['F258'].value - ws['F257'].value,
                         ws['F259'].value - ws['F257'].value]
            tie_info = [[label_1, label_2], False, [False, False], False, [False, False]]
            tie_info[1] = True if ws['E97'].value else False
            tie_info[2][0] = True if ws['N90'].value else False
            tie_info[2][1] = True if ws['N91'].value else False
            tie_info[3] = True if ws['AA97'].value else False
            tie_info[4][0] = True if ws['R90'].value else False
            tie_info[4][1] = True if ws['R91'].value else False
            bars_info['info'].append(dict(zip(bars_keys, [label, case, side, order, left_cut, right_cut, tie_info])))
        # Bottom long bar
        if ws['O135'].value != 0 or ws['O136'].value != 0:
            label = ""
            label_1 = str(ws['O135'].value) + '%%C' + ws['Q135'].value
            label_2 = str(ws['O136'].value) + '%%C' + ws['Q136'].value
            if ws['O135'].value != 0:
                label = label + label_1 + (' + ' if ws['O136'].value != 0 else '')
            if ws['O136'].value != 0:
                label = label + label_2
            case = 0
            side = -1
            order = 0
            left_cut = [ws['F204'].value - ws['F219'].value,
                        ws['F205'].value - ws['F219'].value]
            right_cut = [ws['F273'].value - ws['F257'].value,
                         ws['F274'].value - ws['F257'].value]
            tie_info = [[label_1, label_2], False, [False, False], False, [False, False]]
            tie_info[1] = True if ws['E125'].value else False
            tie_info[2][0] = True if ws['N135'].value else False
            tie_info[2][1] = True if ws['N136'].value else False
            tie_info[3] = True if ws['AA125'].value else False
            tie_info[4][0] = True if ws['R135'].value else False
            tie_info[4][1] = True if ws['R136'].value else False
            bars_info['info'].append(dict(zip(bars_keys, [label, case, side, order, left_cut, right_cut, tie_info])))
        # Top left first order bar
        if ws['J101'].value != 0 or ws['J103'].value != 0:
            label = ""
            label_1 = str(ws['J101'].value) + '%%C' + ws['L101'].value
            label_2 = str(ws['J103'].value) + '%%C' + ws['L103'].value
            if ws['J101'].value != 0:
                label = label + label_1 + (' + ' if ws['J103'].value != 0 else '')
            if ws['J103'].value != 0:
                label = label + label_2
            case = 1
            side = 1
            order = 1
            left_cut = [ws['F194'].value - ws['F219'].value,
                        ws['F195'].value - ws['F219'].value]
            right_cut = [ws['F220'].value - ws['F219'].value,
                         ws['F220'].value - ws['F219'].value]
            tie_info = [[label_1, label_2], False, [False, False]]
            tie_info[1] = True if ws['E99'].value else False
            tie_info[2][0] = True if ws['I101'].value else False
            tie_info[2][1] = True if ws['I103'].value else False
            bars_info['quantity']['top_left'] += 1
            bars_info['info'].append(dict(zip(bars_keys, [label, case, side, order, left_cut, right_cut, tie_info])))
        # Top left second order bar
        if ws['F103'].value != 0 or ws['F105'].value != 0:
            label = ""
            label_1 = str(ws['F103'].value) + '%%C' + ws['H103'].value
            label_2 = str(ws['F105'].value) + '%%C' + ws['H105'].value
            if ws['F103'].value != 0:
                label = label + label_1 + (' + ' if ws['F105'].value != 0 else '')
            if ws['F105'].value != 0:
                label = label + label_2
            case = 1
            side = 1
            order = 2
            left_cut = [ws['F199'].value - ws['F219'].value,
                        ws['F200'].value - ws['F219'].value]
            right_cut = [ws['F223'].value - ws['F219'].value,
                         ws['F223'].value - ws['F219'].value]
            tie_info = [[label_1, label_2], False, [False, False]]
            tie_info[1] = True if ws['E101'].value else False
            tie_info[2][0] = True if ws['E103'].value else False
            tie_info[2][1] = True if ws['E105'].value else False
            bars_info['quantity']['top_left'] += 1
            bars_info['info'].append(dict(zip(bars_keys, [label, case, side, order, left_cut, right_cut, tie_info])))
        # Bottom left first order bar
        if ws['J118'].value != 0 or ws['J120'].value != 0:
            label = ""
            label_1 = str(ws['J118'].value) + '%%C' + ws['L118'].value
            label_2 = str(ws['J120'].value) + '%%C' + ws['L120'].value
            if ws['J118'].value != 0:
                label = label + label_1 + (' + ' if ws['J120'].value != 0 else '')
            if ws['J120'].value != 0:
                label = label + label_2
            case = 1
            side = -1
            order = 1
            left_cut = [ws['F209'].value - ws['F219'].value,
                        ws['F210'].value - ws['F219'].value]
            right_cut = [ws['F226'].value - ws['F219'].value,
                         ws['F226'].value - ws['F219'].value]
            tie_info = [[label_1, label_2], False, [False, False]]
            tie_info[1] = True if ws['E123'].value else False
            tie_info[2][0] = True if ws['I118'].value else False
            tie_info[2][1] = True if ws['I120'].value else False
            bars_info['quantity']['bottom_left'] += 1
            bars_info['info'].append(dict(zip(bars_keys, [label, case, side, order, left_cut, right_cut, tie_info])))
        # Bottom left second order bar
        if ws['F116'].value != 0 or ws['F118'].value != 0:
            label = ""
            label_1 = str(ws['F116'].value) + '%%C' + ws['H116'].value
            label_2 = str(ws['F118'].value) + '%%C' + ws['H118'].value
            if ws['F116'].value != 0:
                label = label + label_1 + (' + ' if ws['F118'].value != 0 else '')
            if ws['F118'].value != 0:
                label = label + label_2
            case = 1
            side = -1
            order = 2
            left_cut = [ws['F214'].value - ws['F219'].value,
                        ws['F215'].value - ws['F219'].value]
            right_cut = [ws['F229'].value - ws['F219'].value,
                         ws['F229'].value - ws['F219'].value]
            tie_info = [[label_1, label_2], False, [False, False]]
            tie_info[1] = True if ws['E121'].value else False
            tie_info[2][0] = True if ws['E116'].value else False
            tie_info[2][1] = True if ws['E118'].value else False
            bars_info['quantity']['bottom_left'] += 1
            bars_info['info'].append(dict(zip(bars_keys, [label, case, side, order, left_cut, right_cut, tie_info])))
        # Bottom central first order bar
        if ws['O116'].value != 0 or ws['O118'].value != 0:
            label = ""
            label_1 = str(ws['O116'].value) + '%%C' + ws['Q116'].value
            label_2 = str(ws['O118'].value) + '%%C' + ws['Q118'].value
            if ws['O116'].value != 0:
                label = label + label_1 + (' + ' if ws['O118'].value != 0 else '')
            if ws['O118'].value != 0:
                label = label + label_2
            case = 2
            side = -1
            order = 1
            left_cut = [ws['F232'].value - ws['F219'].value,
                        ws['F232'].value - ws['F219'].value]
            right_cut = [ws['F251'].value - ws['F257'].value,
                         ws['F251'].value - ws['F257'].value]
            tie_info = [[label_1, label_2], False, False]
            bars_info['quantity']['bottom_center'] += 1
            bars_info['info'].append(dict(zip(bars_keys, [label, case, side, order, left_cut, right_cut, tie_info])))
        # Bottom central second order bar
        if ws['O108'].value != 0 or ws['O110'].value != 0:
            label = ""
            label_1 = str(ws['O108'].value) + '%%C' + ws['Q108'].value
            label_2 = str(ws['O110'].value) + '%%C' + ws['Q110'].value
            if ws['O108'].value != 0:
                label = label + label_1 + (' + ' if ws['O110'].value != 0 else '')
            if ws['O110'].value != 0:
                label = label + label_2
            case = 2
            side = -1
            order = 2
            left_cut = [ws['F235'].value - ws['F219'].value,
                        ws['F235'].value - ws['F219'].value]
            right_cut = [ws['F254'].value - ws['F257'].value,
                         ws['F254'].value - ws['F257'].value]
            tie_info = [[label_1, label_2], False, False]
            bars_info['quantity']['bottom_center'] += 1
            bars_info['info'].append(dict(zip(bars_keys, [label, case, side, order, left_cut, right_cut, tie_info])))
        # Top right first order bar
        if ws['T101'].value != 0 or ws['T103'].value != 0:
            label = ""
            label_1 = str(ws['T101'].value) + '%%C' + ws['V101'].value
            label_2 = str(ws['T103'].value) + '%%C' + ws['V103'].value
            if ws['T101'].value != 0:
                label = label + label_1 + (' + ' if ws['T103'].value != 0 else '')
            if ws['T103'].value != 0:
                label = label + label_2
            case = 3
            side = 1
            order = 1
            left_cut = [ws['F239'].value - ws['F257'].value,
                        ws['F239'].value - ws['F257'].value]
            right_cut = [ws['F263'].value - ws['F257'].value,
                         ws['F264'].value - ws['F257'].value]
            tie_info = [[label_1, label_2], False, [False, False]]
            tie_info[1] = True if ws['AA99'].value else False
            tie_info[2][0] = True if ws['W101'].value else False
            tie_info[2][1] = True if ws['W103'].value else False
            bars_info['quantity']['top_right'] += 1
            bars_info['info'].append(dict(zip(bars_keys, [label, case, side, order, left_cut, right_cut, tie_info])))
        # Top right second order bar
        if ws['X103'].value != 0 or ws['X105'].value != 0:
            label = ""
            label_1 = str(ws['X103'].value) + '%%C' + ws['Z103'].value
            label_2 = str(ws['X105'].value) + '%%C' + ws['Z105'].value
            if ws['X103'].value != 0:
                label = label + label_1 + (' + ' if ws['X105'].value != 0 else '')
            if ws['X105'].value != 0:
                label = label + label_2
            case = 3
            side = 1
            order = 2
            left_cut = [ws['F242'].value - ws['F257'].value,
                        ws['F242'].value - ws['F257'].value]
            right_cut = [ws['F268'].value - ws['F257'].value,
                         ws['F269'].value - ws['F257'].value]
            tie_info = [[label_1, label_2], False, [False, False]]
            tie_info[1] = True if ws['AA101'].value else False
            tie_info[2][0] = True if ws['AA103'].value else False
            tie_info[2][1] = True if ws['AA105'].value else False
            bars_info['quantity']['top_right'] += 1
            bars_info['info'].append(dict(zip(bars_keys, [label, case, side, order, left_cut, right_cut, tie_info])))
        # Bottom right first order bar
        if ws['T118'].value != 0 or ws['T120'].value != 0:
            label = ""
            label_1 = str(ws['T118'].value) + '%%C' + ws['V118'].value
            label_2 = str(ws['T120'].value) + '%%C' + ws['V120'].value
            if ws['T118'].value != 0:
                label = label + label_1 + (' + ' if ws['T120'].value != 0 else '')
            if ws['T120'].value != 0:
                label = label + label_2
            case = 3
            side = -1
            order = 1
            left_cut = [ws['F245'].value - ws['F257'].value,
                        ws['F245'].value - ws['F257'].value]
            right_cut = [ws['F278'].value - ws['F257'].value,
                         ws['F279'].value - ws['F257'].value]
            tie_info = [[label_1, label_2], False, [False, False]]
            tie_info[1] = True if ws['AA123'].value else False
            tie_info[2][0] = True if ws['W118'].value else False
            tie_info[2][1] = True if ws['W120'].value else False
            bars_info['quantity']['bottom_right'] += 1
            bars_info['info'].append(dict(zip(bars_keys, [label, case, side, order, left_cut, right_cut, tie_info])))
        # Bottom right second order bar
        if ws['X116'].value != 0 or ws['X118'].value != 0:
            label = ""
            label_1 = str(ws['X116'].value) + '%%C' + ws['Z116'].value
            label_2 = str(ws['X118'].value) + '%%C' + ws['Z118'].value
            if ws['X116'].value != 0:
                label = label + label_1 + (' + ' if ws['X118'].value != 0 else '')
            if ws['X118'].value != 0:
                label = label + label_2
            case = 3
            side = -1
            order = 2
            left_cut = [ws['F248'].value - ws['F257'].value,
                        ws['F248'].value - ws['F257'].value]
            right_cut = [ws['F283'].value - ws['F257'].value,
                         ws['F284'].value - ws['F257'].value]
            tie_info = [[label_1, label_2], False, [False, False]]
            tie_info[1] = True if ws['AA121'].value else False
            tie_info[2][0] = True if ws['AA116'].value else False
            tie_info[2][1] = True if ws['AA118'].value else False
            bars_info['quantity']['bottom_right'] += 1
            bars_info['info'].append(dict(zip(bars_keys, [label, case, side, order, left_cut, right_cut, tie_info])))

        stirrups_info = {'differentiate': ws['I414'].value, 'diameters': {
            'l2r_diam': "%%C" + str(ws['L416'].value),
            'r2l_diam': "%%C" + str(ws['L424'].value)},
            'quantity': {
            'l2r_two_legged': ws['I416'].value,
            'l2r_single_legged': ws['J416'].value,
            'r2l_two_legged': ws['I424'].value,
            'r2l_single_legged': ws['J424'].value},
            'text': "",
            'info': []}

        stirrups_text = str(ws['I416'].value) + ' (est. rect.) '
        stirrups_text = stirrups_text + ('+ ' + str(ws['J416'].value) + ' (gancho) ' if ws['J416'].value != 0 else '')
        stirrups_text = stirrups_text + '%%C' + ws['L416'].value + ': ' + ws['M431'].value
        if not stirrups_info['differentiate']:
            stirrups_text = stirrups_text + ' c/ext.'
        else:
            stirrups_text = stirrups_text + ' ----->    <----- '
            stirrups_text = stirrups_text + str(ws['I424'].value) + ' (est. rect.) '
            stirrups_text = stirrups_text + (
                '+ ' + str(ws['J424'].value) + ' (gancho) ' if ws['J424'].value != 0 else '')
            stirrups_text = stirrups_text + '%%C' + ws['L424'].value + ': ' + ws['U431'].value
        stirrups_info['text'] = stirrups_text

        stirrups_keys = ['side', 'quantity', 'spacing']
        l2r_row = 417
        while True:
            quantity = ws.cell(l2r_row, 14).value
            spacing = ws.cell(l2r_row, 16).value
            stirrups_info['info'].append(dict(zip(stirrups_keys, [0, quantity, spacing])))
            if ws.cell(l2r_row, 13).value == 1:
                break
            l2r_row += 1
        r2l_row = 425
        while True:
            quantity = ws.cell(r2l_row, 14).value
            spacing = ws.cell(r2l_row, 16).value
            stirrups_info['info'].append(dict(zip(stirrups_keys, [1, quantity, spacing])))
            if ws.cell(r2l_row, 13).value == 1:
                break
            r2l_row += 1
        return dict(zip(span_keys, [span_name, ls_info, rs_info, free_length, width, height, bars_info, stirrups_info]))


if __name__ == '__main__':
    ast = Assistant()
